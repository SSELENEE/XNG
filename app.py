
import json
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote_plus

import numpy as np
import requests

try:
    import shapefile  # pyshp: pure-python shapefile reader
    from shapely.geometry import shape as shapely_shape, Polygon
except Exception:
    shapefile = None
    shapely_shape = None
    Polygon = None
import streamlit as st
from bs4 import BeautifulSoup
from PIL import Image

try:
    import feedparser
except Exception:
    feedparser = None

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
IMG_DIR = ROOT / "images"
REPORT_DIR = ROOT / "reports"

for d in [DATA_DIR, IMG_DIR, REPORT_DIR]:
    d.mkdir(exist_ok=True)

MODULES = ["NOAA 6-10D", "NOAA 8-14D", "CME NG Structure", "News Scan"]
WEEKLY_MODULE = "Weekly Storage"
MONTHLY_MODULE = "Monthly STEO"
BIAS_OPTIONS = ["bearish", "slightly bearish", "neutral", "slightly bullish", "bullish"]

REGION_BOXES = {
    "West": (0.04, 0.23, 0.27, 0.78),
    "Mountain": (0.23, 0.24, 0.43, 0.76),
    "Central": (0.40, 0.25, 0.62, 0.77),
    "Midwest": (0.55, 0.18, 0.75, 0.58),
    "South": (0.50, 0.58, 0.78, 0.86),
    "East": (0.70, 0.24, 0.93, 0.72),
    "Northeast": (0.76, 0.15, 0.95, 0.36),
}
GAS_WEIGHTS = {"West": 0.35, "Mountain": 0.45, "Central": 1.00, "Midwest": 1.10, "South": 1.15, "East": 1.10, "Northeast": 0.75}

BEARISH_TERMS = {
    "glut": 3, "inventory glut": 4, "storage surplus": 4, "oversupply": 4, "build": 2, "injection": 2,
    "mild weather": 4, "warmer weather": 2, "cooler weather": 2, "weak demand": 4, "lower demand": 3,
    "slides": 2, "falls": 2, "extends losses": 3, "bearish": 4, "production strong": 3, "record production": 3,
}
BULLISH_TERMS = {
    "tight supply": 4, "supply shock": 4, "cold": 2, "heat": 2, "hot weather": 4, "strong demand": 4,
    "lng demand": 3, "exports rise": 3, "outage": 3, "pipeline disruption": 4, "withdrawal": 2,
    "rallies": 2, "surges": 3, "bullish": 4, "iran war": 1, "geopolitical": 1, "global supply": 2,
}


def slugify(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")



def choose_final_values(mod: dict, score_value, selected_bias, summary_value, reason_value, notes_value) -> dict:
    """
    Prevent stale Streamlit widgets from overwriting auto_result.
    If fields are still default 0/neutral/empty but auto_result exists, use auto_result.
    """
    ar = mod.get("auto_result") or {}
    visible_is_default = (
        int(score_value) == 0
        and selected_bias == "neutral"
        and not str(summary_value).strip()
        and ar
    )
    if visible_is_default:
        return {
            "score": int(ar.get("score", 0)),
            "bias": ar.get("bias", "neutral"),
            "summary": ar.get("summary", ""),
            "reason": ar.get("reason", ""),
            "raw_notes": notes_value,
        }
    return {
        "score": score_value,
        "bias": selected_bias,
        "summary": summary_value,
        "reason": reason_value,
        "raw_notes": notes_value,
    }


def report_values(mod: dict) -> dict:
    """
    Report source of truth:
    prefer latest auto_result if module fields are stale/default.
    """
    ar = mod.get("auto_result") or {}
    score = mod.get("score", 0)
    bias = mod.get("bias", "neutral")
    summary = mod.get("summary", "")
    reason = mod.get("reason", "")

    if ar and (score == 0 and bias == "neutral" and not summary):
        score = ar.get("score", score)
        bias = ar.get("bias", bias)
        summary = ar.get("summary", summary)
        reason = ar.get("reason", reason)

    return {
        "score": score,
        "bias": bias,
        "summary": summary,
        "reason": reason,
        "auto_result": ar,
    }


def day_file(day: str) -> Path:
    return DATA_DIR / f"{day}.json"


def make_module() -> dict:
    return {"completed": False, "confirmed_at": None, "image_path": None, "shapefile_set": {}, "source_path": None, "score": 0, "bias": "neutral", "summary": "", "reason": "", "raw_notes": "", "auto_result": {}, "field_version": 0}


def make_new_day(day: str) -> dict:
    return {"date": day, "created_at": datetime.now().isoformat(), "modules": {m: make_module() for m in MODULES}, "weekly": {WEEKLY_MODULE: make_weekly_storage()}, "monthly": {MONTHLY_MODULE: make_monthly_steo()}, "daily_report": ""}


def ensure_cme_structure(mod: dict) -> None:
    mod.setdefault("cme_timeframes", {})
    for tf in ["1H", "4H", "1D"]:
        mod["cme_timeframes"].setdefault(tf, {"image_path": None, "score": 0, "bias": "neutral", "summary": "", "reason": "", "auto_result": {}})


def ensure_news(mod: dict) -> None:
    mod.setdefault("news_items", [])
    mod.setdefault("news_query", "US natural gas news")


def make_weekly_storage() -> dict:
    return {
        "completed": False,
        "confirmed_at": None,
        "week_ending": "",
        "actual": 0.0,
        "expected": 0.0,
        "total_storage": 0.0,
        "five_year_avg": 0.0,
        "last_year": 0.0,
        "surprise": 0.0,
        "vs_5y": 0.0,
        "vs_last_year": 0.0,
        "score": 0,
        "bias": "neutral",
        "summary": "",
        "reason": "",
        "raw_notes": "",
        "field_version": 0,
        "auto_result": {},
        "image_path": None,
        "extracted_text": "",
    }


def ensure_weekly_storage(data: dict) -> None:
    data.setdefault("weekly", {})
    data["weekly"].setdefault(WEEKLY_MODULE, make_weekly_storage())
    default = make_weekly_storage()
    for k, v in default.items():
        data["weekly"][WEEKLY_MODULE].setdefault(k, v)


def make_monthly_steo() -> dict:
    return {
        "completed": False,
        "confirmed_at": None,
        "release_month": "",
        "source_path": None,
        "text_source_path": None,
        "score": 0,
        "bias": "neutral",
        "summary": "",
        "reason": "",
        "raw_notes": "",
        "field_version": 0,
        "auto_result": {},
        "metrics": {},
    }


def ensure_monthly_steo(data: dict) -> None:
    data.setdefault("monthly", {})
    data["monthly"].setdefault(MONTHLY_MODULE, make_monthly_steo())
    default = make_monthly_steo()
    for k, v in default.items():
        data["monthly"][MONTHLY_MODULE].setdefault(k, v)


def safe_float(v):
    try:
        if v is None:
            return None
        if isinstance(v, str):
            v = v.replace(",", "").replace("$", "").replace("%", "").strip()
            if v in ["", "-", "—", "n/a", "NA"]:
                return None
        return float(v)
    except Exception:
        return None


def final_bias_from_score(score: float) -> str:
    if score >= 1.5:
        return "bullish"
    if score >= 0.5:
        return "slightly bullish"
    if score <= -1.5:
        return "bearish"
    if score <= -0.5:
        return "slightly bearish"
    return "neutral"


def confidence_from_score(score: float) -> str:
    strength = abs(score)
    if strength >= 1.5:
        return "high"
    if strength >= 0.75:
        return "medium"
    if strength >= 0.35:
        return "low"
    return "very low"


def scan_workbook_for_terms(path: Path, terms: list[str], max_hits: int = 80) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    hits = []
    terms_l = [t.lower() for t in terms]

    for ws in wb.worksheets:
        if ws is None:
            continue

        for row in ws.iter_rows():
            if row is None:
                continue

            vals = []
            for c in row:
                try:
                    vals.append(c.value if c is not None else None)
                except Exception:
                    vals.append(None)

            joined = " | ".join("" if v is None else str(v) for v in vals)
            jl = joined.lower()

            if any(t in jl for t in terms_l):
                hits.append({
                    "sheet": ws.title,
                    "row": getattr(row[0], "row", None) if row else None,
                    "text": joined[:1000],
                    "values": [None if v is None else str(v) for v in vals[:30]],
                })

                if len(hits) >= max_hits:
                    return hits

    return hits

def extract_number_after_phrase(text: str, phrase: str, window: int = 260):
    idx = text.lower().find(phrase.lower())
    if idx == -1:
        return None
    chunk = text[idx:idx + window]
    m = re.search(r"(-?\d+(?:\.\d+)?)", chunk.replace(",", ""))
    return float(m.group(1)) if m else None


def parse_steo_text_narrative(text: str) -> dict:
    clean = " ".join(text.replace(",", "").split())
    low = clean.lower()

    out = {}

    # Pull direct narrative values when visible in STEO text
    m = re.search(r"2026 LNG exports will total\s+([0-9.]+)\s+Bcf/d", clean, flags=re.I)
    if m:
        out["lng_exports_2026"] = float(m.group(1))
    m = re.search(r"2027 exports will total\s+([0-9.]+)\s+Bcf/d", clean, flags=re.I)
    if m:
        out["lng_exports_2027"] = float(m.group(1))

    m = re.search(r"Henry Hub prices in 2Q26 and 3Q26.*?averaging about \$?([0-9.]+)", clean, flags=re.I)
    if m:
        out["henry_hub_summer"] = float(m.group(1))

    m = re.search(r"inventories ended .*? at about\s+([0-9]+)\s+Bcf.*?([0-9]+)% above the five-year average", clean, flags=re.I)
    if m:
        out["ending_storage_bcf"] = float(m.group(1))
        out["ending_storage_vs_5y_pct"] = float(m.group(2))

    m = re.search(r"end October at\s+([0-9]+)\s+Bcf,\s+([0-9]+)% more than the five-year average", clean, flags=re.I)
    if m:
        out["oct_storage_bcf"] = float(m.group(1))
        out["oct_storage_vs_5y_pct"] = float(m.group(2))

    m = re.search(r"marketed natural gas production to increase\s+([0-9]+)% in 2026 and\s+([0-9]+)% in 2027", clean, flags=re.I)
    if m:
        out["production_growth_2026_pct"] = float(m.group(1))
        out["production_growth_2027_pct"] = float(m.group(2))

    m = re.search(r"average about\s+([0-9]+)\s+cooling degree days.*?([0-9]+)% more CDDs than in 2025", clean, flags=re.I)
    if m:
        out["cdd_2026"] = float(m.group(1))
        out["cdd_vs_2025_pct"] = float(m.group(2))

    # key narrative flags
    out["narrative_flags"] = {
        "storage_injections_above_typical": "more natural gas will be injected into storage than is typical" in low or "storage injections to outpace the five-year average" in low,
        "lng_exports_high": "near maximum capacity" in low or "near-peak capacity" in low,
        "production_growth": "production to increase" in low or "natural gas production growth" in low,
        "strong_power_demand": "cooling needs increase" in low or "more cdds" in low,
    }
    return out


def analyze_monthly_steo_metrics(metrics: dict) -> dict:
    """
    Monthly STEO = structural/regime background, not execution trigger.
    Bearish factors: storage above avg, injections above typical, production growth, HH forecast lower/soft.
    Bullish factors: LNG exports high/rising, power burn/CDD above normal, storage below avg.
    """
    score_float = 0.0
    reasons = []

    oct_vs = safe_float(metrics.get("oct_storage_vs_5y_pct"))
    end_vs = safe_float(metrics.get("ending_storage_vs_5y_pct"))
    prod_growth = safe_float(metrics.get("production_growth_2026_pct"))
    lng_2026 = safe_float(metrics.get("lng_exports_2026"))
    cdd_pct = safe_float(metrics.get("cdd_vs_2025_pct"))
    hh_summer = safe_float(metrics.get("henry_hub_summer"))

    flags = metrics.get("narrative_flags", {}) or {}

    if oct_vs is not None:
        if oct_vs >= 5:
            score_float -= 1.2
            reasons.append(f"End-October storage projected {oct_vs:.1f}% above 5Y average: bearish structural inventory backdrop.")
        elif oct_vs > 0:
            score_float -= 0.6
            reasons.append(f"End-October storage projected {oct_vs:.1f}% above 5Y average: slightly bearish inventory backdrop.")
        elif oct_vs <= -5:
            score_float += 1.2
            reasons.append(f"End-October storage projected {abs(oct_vs):.1f}% below 5Y average: bullish inventory backdrop.")
        elif oct_vs < 0:
            score_float += 0.6
            reasons.append(f"End-October storage projected {abs(oct_vs):.1f}% below 5Y average: slightly bullish inventory backdrop.")

    if end_vs is not None and end_vs > 0:
        score_float -= 0.4
        reasons.append(f"Withdrawal-season ending storage {end_vs:.1f}% above 5Y average.")
    elif end_vs is not None and end_vs < 0:
        score_float += 0.4
        reasons.append(f"Withdrawal-season ending storage {abs(end_vs):.1f}% below 5Y average.")

    if prod_growth is not None:
        if prod_growth >= 2:
            score_float -= 0.6
            reasons.append(f"Marketed gas production expected to grow {prod_growth:.1f}% in 2026: supply pressure.")
        elif prod_growth <= -1:
            score_float += 0.6
            reasons.append(f"Marketed gas production expected to fall {abs(prod_growth):.1f}% in 2026: supply tightening.")

    if lng_2026 is not None:
        if lng_2026 >= 17:
            score_float += 0.7
            reasons.append(f"LNG exports projected around {lng_2026:.1f} Bcf/d in 2026: bullish demand pull / export floor.")
        elif lng_2026 <= 15:
            score_float -= 0.4
            reasons.append(f"LNG exports projected around {lng_2026:.1f} Bcf/d: weaker export pull.")

    if cdd_pct is not None:
        if cdd_pct >= 4:
            score_float += 0.5
            reasons.append(f"CDD forecast {cdd_pct:.1f}% above 2025: supportive power-burn demand.")
        elif cdd_pct <= -3:
            score_float -= 0.5
            reasons.append(f"CDD forecast {abs(cdd_pct):.1f}% below 2025: weaker cooling demand.")

    if hh_summer is not None:
        if hh_summer < 3.25:
            score_float -= 0.3
            reasons.append(f"Henry Hub summer price forecast near ${hh_summer:.2f}: market expects contained prices.")
        elif hh_summer > 4.0:
            score_float += 0.3
            reasons.append(f"Henry Hub summer price forecast near ${hh_summer:.2f}: market expects tighter balance.")

    if flags.get("storage_injections_above_typical"):
        score_float -= 0.5
        reasons.append("Narrative says injections should be above typical levels.")
    if flags.get("lng_exports_high"):
        score_float += 0.3
        reasons.append("Narrative says LNG export utilization is high / near peak.")
    if flags.get("strong_power_demand"):
        score_float += 0.3
        reasons.append("Narrative indicates stronger cooling/power demand.")

    if score_float >= 1.5:
        score, bias = 2, "bullish"
    elif score_float >= 0.5:
        score, bias = 1, "slightly bullish"
    elif score_float <= -1.5:
        score, bias = -2, "bearish"
    elif score_float <= -0.5:
        score, bias = -1, "slightly bearish"
    else:
        score, bias = 0, "neutral"

    summary_bits = []
    for k in ["henry_hub_summer", "lng_exports_2026", "ending_storage_bcf", "ending_storage_vs_5y_pct", "oct_storage_bcf", "oct_storage_vs_5y_pct", "production_growth_2026_pct", "cdd_2026", "cdd_vs_2025_pct"]:
        if metrics.get(k) is not None:
            summary_bits.append(f"{k}={metrics[k]}")
    summary = "; ".join(summary_bits) if summary_bits else "Monthly STEO parsed, but key natural-gas metrics are sparse. Review manually."

    return {
        "score": score,
        "bias": bias,
        "summary": summary,
        "reason": " ".join(reasons) if reasons else "Monthly STEO is broadly neutral or insufficiently parsed.",
        "metrics": metrics,
        "score_float": round(score_float, 2),
        "generated_at": datetime.now().isoformat(),
        "engine": "v0.8 monthly STEO regime engine",
    }




def steo_release_label_from_result(res: dict) -> str:
    raw = res.get("release_date") or res.get("fetched_at") or ""
    if not raw:
        return ""
    # Prefer human-readable release date when available, e.g. "April 7, 2026"
    return str(raw).strip()

def fetch_latest_steo_page() -> dict:
    """
    Fetch latest EIA STEO page directly from EIA.
    This avoids PDF OCR and avoids fragile workbook scanning.
    """
    url = "https://www.eia.gov/outlooks/steo/"
    headers = {"User-Agent": "Mozilla/5.0 XNGResearchDesk/0.9"}
    r = requests.get(url, headers=headers, timeout=20)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")
    page_text = soup.get_text(" ", strip=True)

    # Try to find release date / next release date from page text
    release_date = ""
    m = re.search(r"Release Date:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})", page_text, flags=re.I)
    if m:
        release_date = m.group(1)

    next_release_date = ""
    m = re.search(r"Next Release Date:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})", page_text, flags=re.I)
    if m:
        next_release_date = m.group(1)

    return {
        "url": url,
        "release_date": release_date,
        "next_release_date": next_release_date,
        "page_text": page_text,
        "fetched_at": datetime.now().isoformat(),
    }


def extract_natural_gas_section_from_steo_text(text: str) -> str:
    """
    Pull the most relevant STEO natural gas / electricity demand / LNG / storage narrative.
    EIA page layout can change, so this uses anchor windows instead of strict DOM assumptions.
    """
    clean = " ".join(text.split())
    low = clean.lower()

    anchors = [
        "natural gas inventories",
        "lng exports",
        "natural gas spot price",
        "henry hub",
        "electricity demand",
        "cooling degree days",
        "natural gas production",
    ]

    chunks = []
    for a in anchors:
        idx = low.find(a)
        if idx != -1:
            start = max(0, idx - 700)
            end = min(len(clean), idx + 2200)
            chunk = clean[start:end]
            if chunk not in chunks:
                chunks.append(chunk)

    if not chunks:
        # fallback: use page text but cap it
        return clean[:12000]

    return "\n\n".join(chunks)[:16000]


def analyze_latest_steo_fetch() -> dict:
    fetched = fetch_latest_steo_page()
    ng_text = extract_natural_gas_section_from_steo_text(fetched["page_text"])
    metrics = parse_steo_text_narrative(ng_text)

    res = analyze_monthly_steo_metrics(metrics)
    res["source"] = "steo_html_fetch"
    res["url"] = fetched["url"]
    res["release_date"] = fetched["release_date"]
    res["next_release_date"] = fetched["next_release_date"]
    res["extracted_text"] = ng_text[:5000]
    res["fetched_at"] = fetched["fetched_at"]
    return res


def extract_steo_from_pdf_text(pdf_text: str) -> dict:
    metrics = parse_steo_text_narrative(pdf_text)
    return analyze_monthly_steo_metrics(metrics)



def _sheet_text(ws, max_rows=80, max_cols=80) -> str:
    rows = []
    for row in ws.iter_rows(max_row=max_rows, max_col=max_cols):
        vals = []
        for c in row:
            try:
                vals.append("" if c.value is None else str(c.value))
            except Exception:
                vals.append("")
        rows.append(" | ".join(vals))
    return "\n".join(rows)


def _find_rows_by_label(ws, labels: list[str], max_rows=160, max_cols=8):
    out = {}
    labels_l = {label: label.lower() for label in labels}
    for row in ws.iter_rows(max_row=max_rows, max_col=max_cols):
        vals = []
        for c in row:
            try:
                vals.append("" if c.value is None else str(c.value))
            except Exception:
                vals.append("")
        joined = " ".join(vals).lower()
        for label, ll in labels_l.items():
            if ll in joined and label not in out:
                out[label] = row[0].row if row and hasattr(row[0], "row") else None
    return out


def _extract_monthly_values_from_row(ws, row_num: int, start_col=3, end_col=90) -> list[float]:
    vals = []
    if not row_num:
        return vals
    for col in range(start_col, end_col + 1):
        v = ws.cell(row=row_num, column=col).value
        fv = safe_float(v)
        if fv is not None:
            vals.append(fv)
    return vals


def _summarize_series(values: list[float]) -> dict:
    if not values:
        return {"first": None, "last": None, "avg": None, "delta": None, "min": None, "max": None, "count": 0}
    first = values[0]
    last = values[-1]
    avg = sum(values) / len(values)
    return {
        "first": round(first, 4),
        "last": round(last, 4),
        "avg": round(avg, 4),
        "delta": round(last - first, 4),
        "min": round(min(values), 4),
        "max": round(max(values), 4),
        "count": len(values),
    }


def extract_steo_table_metrics(path: Path) -> dict:
    """
    Targeted STEO workbook extractor for NG monthly regime.
    Table 5a: supply / consumption / inventories.
    Table 5b: Henry Hub / regional prices.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    table5a = None
    table5b = None

    for ws in wb.worksheets:
        txt = _sheet_text(ws, max_rows=20, max_cols=12).lower()
        if ("natural gas supply" in txt and "consumption" in txt and "inventories" in txt) or "table 5a" in txt:
            table5a = ws
        if ("u.s. regional natural gas prices" in txt) or "table 5b" in txt:
            table5b = ws

    metrics = {
        "table_extractor": "v0.9.1 targeted STEO table parser",
        "table5a_found": table5a.title if table5a else None,
        "table5b_found": table5b.title if table5b else None,
    }
    evidence = {}

    if table5b is not None:
        rows = _find_rows_by_label(table5b, ["Henry Hub spot price"], max_rows=80, max_cols=12)
        hh_row = rows.get("Henry Hub spot price")
        hh_vals = _extract_monthly_values_from_row(table5b, hh_row, start_col=3, end_col=90)
        metrics["henry_hub_series"] = _summarize_series(hh_vals)
        evidence["henry_hub_row"] = hh_row
        if hh_vals:
            near = hh_vals[:min(len(hh_vals), 6)]
            late = hh_vals[-min(len(hh_vals), 6):]
            metrics["henry_hub_near_avg"] = round(sum(near) / len(near), 4)
            metrics["henry_hub_late_avg"] = round(sum(late) / len(late), 4)
            metrics["henry_hub_curve_delta"] = round(metrics["henry_hub_late_avg"] - metrics["henry_hub_near_avg"], 4)

    if table5a is not None:
        rows = _find_rows_by_label(table5a, [
            "U.S. total marketed natural gas production",
            "Total consumption",
            "Electric power",
            "LNG gross exports",
            "End-of-period working natural gas inventories",
            "United States total",
        ], max_rows=170, max_cols=12)

        prod_row = rows.get("U.S. total marketed natural gas production")
        prod_vals = _extract_monthly_values_from_row(table5a, prod_row, start_col=3, end_col=90)
        metrics["production_series"] = _summarize_series(prod_vals)
        evidence["production_row"] = prod_row

        cons_row = rows.get("Total consumption")
        cons_vals = _extract_monthly_values_from_row(table5a, cons_row, start_col=3, end_col=90)
        metrics["consumption_series"] = _summarize_series(cons_vals)
        evidence["consumption_row"] = cons_row

        power_row = rows.get("Electric power")
        power_vals = _extract_monthly_values_from_row(table5a, power_row, start_col=3, end_col=90)
        metrics["power_burn_series"] = _summarize_series(power_vals)
        evidence["power_burn_row"] = power_row

        lng_row = rows.get("LNG gross exports")
        lng_vals = _extract_monthly_values_from_row(table5a, lng_row, start_col=3, end_col=90)
        metrics["lng_exports_series"] = _summarize_series(lng_vals)
        evidence["lng_exports_row"] = lng_row

        inv_header = rows.get("End-of-period working natural gas inventories")
        inv_row = None
        if inv_header:
            for r in range(inv_header, min(inv_header + 12, table5a.max_row) + 1):
                row_txt = " ".join("" if table5a.cell(r, c).value is None else str(table5a.cell(r, c).value) for c in range(1, 8)).lower()
                if "united states total" in row_txt:
                    inv_row = r
                    break
        if inv_row is None:
            inv_row = rows.get("United States total")

        inv_vals = _extract_monthly_values_from_row(table5a, inv_row, start_col=3, end_col=90)
        metrics["inventory_series"] = _summarize_series(inv_vals)
        evidence["inventory_row"] = inv_row

        if prod_vals and cons_vals:
            n = min(len(prod_vals), len(cons_vals))
            metrics["prod_minus_consumption_series"] = _summarize_series([prod_vals[i] - cons_vals[i] for i in range(n)])

        if lng_vals and prod_vals:
            n = min(len(lng_vals), len(prod_vals))
            shares = [(lng_vals[i] / prod_vals[i]) * 100 for i in range(n) if prod_vals[i] != 0]
            metrics["lng_share_of_production_pct_series"] = _summarize_series(shares)

    metrics["evidence_rows"] = evidence
    return metrics


def analyze_steo_table_metrics(metrics: dict, narrative_metrics: dict | None = None) -> dict:
    narrative_metrics = narrative_metrics or {}
    merged = {**narrative_metrics, **metrics}

    score_float = 0.0
    reasons = []

    hh_delta = safe_float(metrics.get("henry_hub_curve_delta"))
    if hh_delta is not None:
        if hh_delta >= 0.5:
            score_float += 0.7
            reasons.append(f"Henry Hub forecast curve rises by {hh_delta:.2f}: tighter forward price path.")
        elif hh_delta <= -0.5:
            score_float -= 0.7
            reasons.append(f"Henry Hub forecast curve falls by {abs(hh_delta):.2f}: looser forward price path.")
        else:
            reasons.append(f"Henry Hub forecast curve is relatively flat ({hh_delta:+.2f}).")

    prod_delta = safe_float((metrics.get("production_series") or {}).get("delta"))
    if prod_delta is not None:
        if prod_delta >= 3:
            score_float -= 0.8
            reasons.append(f"Marketed production rises by {prod_delta:.1f} Bcf/d across forecast path: supply pressure.")
        elif prod_delta <= -3:
            score_float += 0.8
            reasons.append(f"Marketed production falls by {abs(prod_delta):.1f} Bcf/d across forecast path: supply tightening.")

    lng_series = metrics.get("lng_exports_series") or {}
    lng_delta = safe_float(lng_series.get("delta"))
    lng_last = safe_float(lng_series.get("last"))
    if lng_delta is not None:
        if lng_delta >= 1:
            score_float += 0.8
            reasons.append(f"LNG exports rise by {lng_delta:.1f} Bcf/d: stronger export demand pull.")
        elif lng_delta <= -1:
            score_float -= 0.6
            reasons.append(f"LNG exports fall by {abs(lng_delta):.1f} Bcf/d: weaker export demand pull.")
    if lng_last is not None and lng_last >= 17:
        score_float += 0.4
        reasons.append(f"LNG exports end near {lng_last:.1f} Bcf/d: high export floor.")

    power_delta = safe_float((metrics.get("power_burn_series") or {}).get("delta"))
    if power_delta is not None:
        if power_delta >= 2:
            score_float += 0.5
            reasons.append(f"Electric power gas burn rises by {power_delta:.1f} Bcf/d: supportive demand.")
        elif power_delta <= -2:
            score_float -= 0.4
            reasons.append(f"Electric power gas burn falls by {abs(power_delta):.1f} Bcf/d: softer demand.")

    inv_series = metrics.get("inventory_series") or {}
    inv_delta = safe_float(inv_series.get("delta"))
    inv_last = safe_float(inv_series.get("last"))
    if inv_delta is not None:
        if inv_delta >= 400:
            score_float -= 1.0
            reasons.append(f"End-period inventories rise by {inv_delta:.0f} Bcf: bearish storage accumulation.")
        elif inv_delta <= -400:
            score_float += 1.0
            reasons.append(f"End-period inventories fall by {abs(inv_delta):.0f} Bcf: bullish storage tightening.")
    if inv_last is not None and inv_last >= 3800:
        score_float -= 0.5
        reasons.append(f"Inventories end near {inv_last:.0f} Bcf: high stock overhang risk.")

    bal_delta = safe_float((metrics.get("prod_minus_consumption_series") or {}).get("delta"))
    if bal_delta is not None:
        if bal_delta >= 3:
            score_float -= 0.5
            reasons.append(f"Production-minus-consumption balance widens by {bal_delta:.1f} Bcf/d: looser domestic balance.")
        elif bal_delta <= -3:
            score_float += 0.5
            reasons.append(f"Production-minus-consumption balance tightens by {abs(bal_delta):.1f} Bcf/d.")

    narrative_res = analyze_monthly_steo_metrics(narrative_metrics) if narrative_metrics else None
    if narrative_res and narrative_res.get("score", 0) != 0:
        score_float += 0.25 * int(narrative_res["score"])
        reasons.append(f"Narrative overlay: {narrative_res.get('bias')} ({narrative_res.get('summary')}).")

    if score_float >= 1.5:
        score, bias = 2, "bullish"
    elif score_float >= 0.5:
        score, bias = 1, "slightly bullish"
    elif score_float <= -1.5:
        score, bias = -2, "bearish"
    elif score_float <= -0.5:
        score, bias = -1, "slightly bearish"
    else:
        score, bias = 0, "neutral"

    summary_parts = []
    for name, value, suffix in [
        ("HH curve delta", hh_delta, ""),
        ("production delta", prod_delta, " Bcf/d"),
        ("LNG exports delta", lng_delta, " Bcf/d"),
        ("power burn delta", power_delta, " Bcf/d"),
        ("inventory delta", inv_delta, " Bcf"),
    ]:
        if value is not None:
            summary_parts.append(f"{name}={value:+.2f}{suffix}")

    return {
        "score": score,
        "bias": bias,
        "summary": "; ".join(summary_parts) if summary_parts else "STEO table metrics extracted, but key series are sparse.",
        "reason": " ".join(reasons) if reasons else "Table-driven STEO balance is neutral or insufficiently parsed.",
        "metrics": merged,
        "score_float": round(score_float, 2),
        "generated_at": datetime.now().isoformat(),
        "engine": "v0.9.1 table-driven monthly STEO regime engine",
    }



def extract_steo_from_xlsx(path: Path) -> dict:
    """
    v0.9.1 table-first STEO workbook extractor.
    Uses Table 5a/5b targeted series when possible, with generic scan as fallback evidence.
    """
    table_metrics = extract_steo_table_metrics(path)

    terms = [
        "Henry Hub",
        "Natural Gas Henry Hub",
        "Dry Natural Gas Production",
        "Marketed Natural Gas Production",
        "U.S. marketed natural gas production",
        "LNG exports",
        "liquefied natural gas",
        "working natural gas in storage",
        "cooling degree",
        "heating degree",
        "Natural Gas",
    ]
    hits = scan_workbook_for_terms(path, terms, max_hits=120)
    text = "\n".join(h["text"] for h in hits)
    narrative_metrics = parse_steo_text_narrative(text)
    table_metrics["xlsx_hits"] = hits[:30]

    res = analyze_steo_table_metrics(table_metrics, narrative_metrics=narrative_metrics)
    res["source"] = "xlsx_table_targeted"
    res["source_path"] = str(path)
    return res


def save_uploaded_source(selected_day: str, label: str, uploaded) -> str:
    safe_label = re.sub(r"[^a-zA-Z0-9]+", "_", label.lower()).strip("_")
    ext = Path(uploaded.name).suffix.lower()
    out = DATA_DIR / f"{selected_day}_{safe_label}{ext}"
    out.write_bytes(uploaded.getbuffer())
    return str(out.relative_to(ROOT))


def analyze_weekly_storage(actual: float, expected: float, total_storage: float, five_year_avg: float, last_year: float) -> dict:
    surprise = actual - expected
    vs_5y = total_storage - five_year_avg
    vs_ly = total_storage - last_year

    score_float = 0.0
    reasons = []

    if surprise >= 15:
        score_float -= 2.0
        reasons.append("Large bearish storage surprise: actual build materially above expectation.")
    elif surprise >= 5:
        score_float -= 1.0
        reasons.append("Bearish storage surprise: actual build above expectation.")
    elif surprise <= -15:
        score_float += 2.0
        reasons.append("Large bullish storage surprise: actual build materially below expectation.")
    elif surprise <= -5:
        score_float += 1.0
        reasons.append("Bullish storage surprise: actual build below expectation.")
    else:
        reasons.append("Storage surprise is near expectation.")

    if vs_5y > 100:
        score_float -= 1.0
        reasons.append("Total storage is materially above 5-year average.")
    elif vs_5y > 0:
        score_float -= 0.5
        reasons.append("Total storage is above 5-year average.")
    elif vs_5y < -100:
        score_float += 1.0
        reasons.append("Total storage is materially below 5-year average.")
    elif vs_5y < 0:
        score_float += 0.5
        reasons.append("Total storage is below 5-year average.")

    if vs_ly > 100:
        score_float -= 0.5
        reasons.append("Total storage is materially above last year.")
    elif vs_ly < -100:
        score_float += 0.5
        reasons.append("Total storage is materially below last year.")

    if score_float >= 1.5:
        score, bias = 2, "bullish"
    elif score_float >= 0.5:
        score, bias = 1, "slightly bullish"
    elif score_float <= -1.5:
        score, bias = -2, "bearish"
    elif score_float <= -0.5:
        score, bias = -1, "slightly bearish"
    else:
        score, bias = 0, "neutral"

    summary = (
        f"Actual={actual:.1f} Bcf, Expected={expected:.1f} Bcf, "
        f"Surprise={surprise:+.1f} Bcf. Total={total_storage:.1f} Bcf, "
        f"vs 5Y={vs_5y:+.1f} Bcf, vs LY={vs_ly:+.1f} Bcf."
    )

    return {
        "score": score,
        "bias": bias,
        "summary": summary,
        "reason": " ".join(reasons),
        "metrics": {
            "actual": actual,
            "expected": expected,
            "surprise": surprise,
            "total_storage": total_storage,
            "five_year_avg": five_year_avg,
            "last_year": last_year,
            "vs_5y": vs_5y,
            "vs_last_year": vs_ly,
            "score_float": round(score_float, 2),
        },
        "generated_at": datetime.now().isoformat(),
        "engine": "v0.7 weekly storage engine",
    }


def normalize_day(data: dict, day: str) -> dict:
    data.setdefault("date", day)
    data.setdefault("modules", {})
    for m in MODULES:
        if m not in data["modules"] or not isinstance(data["modules"][m], dict):
            data["modules"][m] = make_module()
        else:
            default = make_module()
            for k, v in default.items():
                data["modules"][m].setdefault(k, v)
        if m == "CME NG Structure":
            ensure_cme_structure(data["modules"][m])
        if m == "News Scan":
            ensure_news(data["modules"][m])
    ensure_weekly_storage(data)
    ensure_monthly_steo(data)
    data.setdefault("daily_report", "")
    return data


def load_day(day: str) -> dict:
    fp = day_file(day)
    if not fp.exists():
        return make_new_day(day)
    try:
        txt = fp.read_text(encoding="utf-8").strip()
        if not txt:
            raise ValueError("empty json")
        return normalize_day(json.loads(txt), day)
    except Exception:
        backup = fp.with_suffix(f".corrupt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        try:
            fp.rename(backup)
        except Exception:
            pass
        return make_new_day(day)


def save_day(day: str, data: dict) -> None:
    data["updated_at"] = datetime.now().isoformat()
    day_file(day).write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")




# =========================
# NOAA SHAPEFILE PRIMARY ENGINE v0.9.6 STREAMLIT CLOUD SAFE + ZIP UPLOAD
# =========================

NOAA_NCDC_REGIONS = [
    "Northwest", "West", "Southwest", "West North Central",
    "East North Central", "Central", "South", "Southeast", "Northeast"
]

NOAA_GAS_WEIGHTS = {
    "Northwest": 0.30,
    "West": 0.40,
    "Southwest": 0.45,
    "West North Central": 0.70,
    "East North Central": 1.15,
    "Central": 1.00,
    "South": 1.10,
    "Southeast": 1.15,
    "Northeast": 0.85,
}

NCDC_REGION_POLYGONS_LONLAT = {
    "Northwest": [(-125, 42), (-125, 49.5), (-111, 49.5), (-111, 42), (-117, 42), (-125, 42)],
    "West": [(-124.8, 32.0), (-124.8, 42.2), (-117.0, 42.2), (-114.0, 36.0), (-114.5, 32.0), (-124.8, 32.0)],
    "Southwest": [(-114.5, 31.2), (-114.0, 42.2), (-102.0, 42.2), (-102.0, 37.0), (-103.0, 37.0), (-103.0, 31.2), (-114.5, 31.2)],
    "West North Central": [(-116.0, 37.0), (-116.0, 49.5), (-94.0, 49.5), (-94.0, 37.0), (-116.0, 37.0)],
    "East North Central": [(-97.5, 36.5), (-97.5, 49.5), (-80.0, 49.5), (-80.0, 38.0), (-84.8, 38.0), (-89.0, 36.5), (-97.5, 36.5)],
    "Central": [(-94.8, 29.0), (-94.8, 39.5), (-80.0, 39.5), (-80.0, 30.0), (-88.0, 29.0), (-94.8, 29.0)],
    "South": [(-106.8, 25.0), (-106.8, 37.0), (-94.0, 37.0), (-94.0, 29.0), (-97.0, 25.0), (-106.8, 25.0)],
    "Southeast": [(-85.5, 24.0), (-85.5, 38.5), (-75.0, 38.5), (-75.0, 24.0), (-85.5, 24.0)],
    "Northeast": [(-80.5, 37.0), (-80.5, 47.8), (-66.5, 47.8), (-66.5, 37.0), (-80.5, 37.0)],
}

def noaa_prob_to_score(cat, prob) -> float:
    c = str(cat or "").strip().lower()
    try:
        p = float(prob)
    except Exception:
        p = 33.0
    mag = max(0.0, min(2.0, (p - 33.0) / 67.0 * 2.0))
    if c.startswith("above"):
        return mag
    if c.startswith("below"):
        return -mag
    return 0.0

def noaa_bias_from_float(score: float) -> str:
    if score >= 1.10:
        return "bullish"
    if score >= 0.35:
        return "slightly bullish"
    if score <= -1.10:
        return "bearish"
    if score <= -0.35:
        return "slightly bearish"
    return "neutral"

def noaa_discrete_score(score: float) -> int:
    if score >= 1.10:
        return 2
    if score >= 0.35:
        return 1
    if score <= -1.10:
        return -2
    if score <= -0.35:
        return -1
    return 0

@st.cache_data(show_spinner=False)
def load_lightweight_ncdc_regions_cached():
    if Polygon is None:
        raise RuntimeError("shapely is required. requirements.txt must include shapely")
    return {name: Polygon(coords) for name, coords in NCDC_REGION_POLYGONS_LONLAT.items()}

def save_uploaded_noaa_shapefile_set(day: str, module: str, uploaded_files: list) -> dict:
    safe_module = slugify(module)
    outdir = DATA_DIR / f"{day}_{safe_module}_shapefile"
    outdir.mkdir(parents=True, exist_ok=True)

    saved = []
    extracted = []
    errors = []

    def safe_write_file(target_dir: Path, filename: str, data: bytes) -> Path:
        clean_name = Path(filename).name
        out = target_dir / clean_name
        out.write_bytes(data)
        return out

    for up in uploaded_files:
        name = Path(up.name).name
        suffix = Path(name).suffix.lower()

        if suffix == ".zip":
            zip_path = safe_write_file(outdir, name, up.getbuffer())
            saved.append(str(zip_path.relative_to(ROOT)))
            try:
                with zipfile.ZipFile(zip_path, "r") as z:
                    for info in z.infolist():
                        if info.is_dir():
                            continue
                        inner_name = Path(info.filename).name
                        if not inner_name:
                            continue
                        inner_lower = inner_name.lower()
                        inner_suffix = Path(inner_name).suffix.lower()
                        allowed = {".shp", ".dbf", ".shx", ".prj", ".cpg", ".sbn", ".sbx", ".xml"}
                        if inner_suffix not in allowed and not inner_lower.endswith(".shp.xml"):
                            continue
                        out = safe_write_file(outdir, inner_name, z.read(info.filename))
                        extracted.append(str(out.relative_to(ROOT)))
            except Exception as e:
                errors.append(f"ZIP extract failed for {name}: {e}")
        else:
            out = safe_write_file(outdir, name, up.getbuffer())
            saved.append(str(out.relative_to(ROOT)))

    shp_files = sorted([p for p in outdir.iterdir() if p.suffix.lower() == ".shp" and not p.name.lower().endswith(".shp.xml")])
    dbf_files = sorted([p for p in outdir.iterdir() if p.suffix.lower() == ".dbf"])
    shx_files = sorted([p for p in outdir.iterdir() if p.suffix.lower() == ".shx"])

    if not shp_files:
        raise RuntimeError("No .shp file found. Upload a ZIP containing .shp/.dbf/.shx or select .shp together with .dbf and .shx.")
    if not dbf_files:
        raise RuntimeError("No .dbf file found. Shapefile attribute table is required.")
    if not shx_files:
        raise RuntimeError("No .shx file found. Shapefile index is required.")

    chosen_shp = None
    for shp in shp_files:
        stem = shp.stem
        if (outdir / f"{stem}.dbf").exists() and (outdir / f"{stem}.shx").exists():
            chosen_shp = shp
            break
    if chosen_shp is None:
        chosen_shp = shp_files[0]

    all_files = sorted([str(p.relative_to(ROOT)) for p in outdir.iterdir() if p.is_file()])

    return {
        "dir": str(outdir.relative_to(ROOT)),
        "shp_path": str(chosen_shp.relative_to(ROOT)),
        "files": all_files,
        "saved_uploads": saved,
        "extracted_files": extracted,
        "mode": "zip" if extracted else "multi-file",
        "errors": errors,
    }

def _read_noaa_records_pyshp(shp_path: Path):
    if shapefile is None or shapely_shape is None:
        raise RuntimeError("pyshp + shapely are required. requirements.txt must include pyshp and shapely")
    reader = shapefile.Reader(str(shp_path))
    fields = [f[0] for f in reader.fields[1:]]
    rows = []
    for sr in reader.iterShapeRecords():
        attrs = dict(zip(fields, sr.record))
        geom = shapely_shape(sr.shape.__geo_interface__)
        rows.append({"attrs": attrs, "geometry": geom})
    return rows, fields

def analyze_noaa_temperature_shapefile(shp_path: Path, module: str) -> dict:
    records, fields = _read_noaa_records_pyshp(shp_path)
    if not records:
        raise RuntimeError("NOAA shapefile contains no records.")

    missing = {"Prob", "Cat"} - set(fields)
    if missing:
        raise RuntimeError(f"NOAA shapefile missing fields: {missing}")

    metadata = {}
    for col in ["Fcst_Date", "Start_Date", "End_Date"]:
        if col in fields:
            metadata[col] = str(records[0]["attrs"].get(col, ""))

    regions = load_lightweight_ncdc_regions_cached()

    noaa_cells = []
    for rec in records:
        attrs = rec["attrs"]
        geom = rec["geometry"]
        cat = attrs.get("Cat")
        prob = attrs.get("Prob")
        sc = noaa_prob_to_score(cat, prob)
        noaa_cells.append({"cat": cat, "prob": prob, "score": sc, "geometry": geom})

    region_rows = []
    total_weighted = 0.0
    total_weight = 0.0

    for region in NOAA_NCDC_REGIONS:
        reg_geom = regions[region]
        reg_area = reg_geom.area
        area_sum = 0.0
        score_area_sum = 0.0
        contributors = []

        for cell in noaa_cells:
            if not reg_geom.intersects(cell["geometry"]):
                continue
            inter = reg_geom.intersection(cell["geometry"])
            if inter.is_empty:
                continue
            a = inter.area
            if a <= 0:
                continue
            score_area_sum += float(cell["score"]) * a
            area_sum += a
            try:
                p = float(cell["prob"])
            except Exception:
                p = 33.0
            contributors.append({
                "cat": str(cell["cat"]),
                "prob": p,
                "cell_score": round(float(cell["score"]), 4),
                "area_share_of_region": round(float(a / reg_area), 6) if reg_area else 0.0,
            })

        raw = score_area_sum / area_sum if area_sum > 0 else 0.0
        wt = NOAA_GAS_WEIGHTS.get(region, 1.0)
        weighted = raw * wt
        total_weighted += weighted
        total_weight += wt

        region_rows.append({
            "region": region,
            "raw_score": round(raw, 4),
            "bias": noaa_bias_from_float(raw),
            "gas_weight": wt,
            "weighted_score": round(weighted, 4),
            "covered_area_share": round(float(area_sum / reg_area), 4) if reg_area else 0.0,
            "contrib_count": len(contributors),
            "contributors": contributors,
        })

    composite = total_weighted / total_weight if total_weight else 0.0
    final_score = noaa_discrete_score(composite)
    bias = noaa_bias_from_float(composite)

    warm_regions = [r["region"] for r in region_rows if r["raw_score"] >= 0.35]
    cool_regions = [r["region"] for r in region_rows if r["raw_score"] <= -0.35]
    mixed_regions = [r["region"] for r in region_rows if -0.35 < r["raw_score"] < 0.35]

    parts = []
    if warm_regions:
        parts.append("Warm/above-normal signal: " + ", ".join(warm_regions))
    if cool_regions:
        parts.append("Cool/below-normal signal: " + ", ".join(cool_regions))
    if not parts:
        parts.append("Mixed/near-normal temperature pattern.")
    if mixed_regions:
        parts.append("Neutral/mixed regions: " + ", ".join(mixed_regions))

    if "bullish" in bias:
        reason = "NOAA shapefile shows area-weighted above-normal temperature pressure in gas-relevant regions. If persistent, this can support power burn / cooling demand."
    elif "bearish" in bias:
        reason = "NOAA shapefile shows area-weighted below-normal temperature pressure in gas-relevant regions. This can suppress cooling demand / power burn."
    else:
        reason = "NOAA shapefile signal is mixed or weak after area-weighting across gas-relevant regions. Treat as no dominant weather edge from this map alone."

    return {
        "score": final_score,
        "bias": bias,
        "summary": " ".join(parts),
        "reason": reason,
        "region_scores": {r["region"]: r["raw_score"] for r in region_rows},
        "region_details": region_rows,
        "composite_raw": round(composite, 4),
        "source_metadata": metadata,
        "analyzer": "v0.9.6 noaa shapefile primary pyshp/shapely + zip upload",
        "generated_at": datetime.now().isoformat(),
        "source_path": str(shp_path),
        "module": module,
    }


def save_uploaded_image(day: str, module: str, uploaded_file) -> str:
    ext = Path(uploaded_file.name).suffix or ".png"
    out = IMG_DIR / f"{day}_{slugify(module)}{ext}"
    with out.open("wb") as f:
        f.write(uploaded_file.getbuffer())
    return str(out.relative_to(ROOT))


def crop_probable_temperature_map(img: Image.Image) -> Image.Image:
    rgb = img.convert("RGB")
    arr = np.array(rgb).astype(np.float32)
    h, w, _ = arr.shape
    y0, y1 = int(h * 0.20), int(h * 0.62)
    section = arr[y0:y1]
    r, g, b = section[:, :, 0], section[:, :, 1], section[:, :, 2]
    maxc, minc = np.maximum.reduce([r, g, b]), np.minimum.reduce([r, g, b])
    mask = ((maxc - minc) > 38) & (maxc > 85)
    ys, xs = np.where(mask)
    if len(xs) < 1000:
        return rgb.crop((int(w * 0.16), int(h * 0.30), int(w * 0.86), int(h * 0.58)))
    pad_x, pad_y = int(w * 0.04), int(h * 0.015)
    x_min, x_max = max(0, int(xs.min()) - pad_x), min(w, int(xs.max()) + pad_x)
    yy_min, yy_max = max(0, y0 + int(ys.min()) - pad_y), min(h, y0 + int(ys.max()) + pad_y)
    return rgb.crop((x_min, yy_min, x_max, yy_max))


def region_temperature_score(arr: np.ndarray) -> float:
    arr = arr.astype(np.float32)
    r, b = arr[:, :, 0], arr[:, :, 2]
    warm = ((r > b + 25) & (r > 95)).sum()
    cool = ((b > r + 25) & (b > 95)).sum()
    return float((warm - cool) / max(arr.shape[0] * arr.shape[1], 1) * 5.0)


def analyze_noaa_temperature_image(image_path: Path, module: str) -> dict:
    img = Image.open(image_path).convert("RGB")
    temp_map = crop_probable_temperature_map(img)
    arr = np.array(temp_map)
    h, w, _ = arr.shape
    region_scores, weighted_sum, weight_sum = {}, 0.0, 0.0
    for region, (x0, y0, x1, y1) in REGION_BOXES.items():
        crop = arr[int(h * y0):int(h * y1), int(w * x0):int(w * x1)]
        score = max(-2.0, min(2.0, region_temperature_score(crop) if crop.size else 0.0))
        region_scores[region] = round(score, 2)
        wt = GAS_WEIGHTS.get(region, 1.0)
        weighted_sum += score * wt
        weight_sum += wt
    composite = weighted_sum / max(weight_sum, 1e-9)
    if composite >= 0.65:
        final_score, bias = 1, "slightly bullish"
    elif composite <= -0.65:
        final_score, bias = -1, "slightly bearish"
    else:
        final_score, bias = 0, "neutral"
    key_avg = sum(region_scores[k] for k in ["Central", "Midwest", "South", "East"]) / 4
    if key_avg >= 1.25:
        final_score, bias = 2, "bullish"
    elif key_avg <= -1.25:
        final_score, bias = -2, "bearish"
    warm_regions = [k for k, v in region_scores.items() if v >= 0.45]
    cool_regions = [k for k, v in region_scores.items() if v <= -0.45]
    parts = []
    if warm_regions: parts.append("Warm/above-normal signal: " + ", ".join(warm_regions))
    if cool_regions: parts.append("Cool/below-normal signal: " + ", ".join(cool_regions))
    if not parts: parts.append("Mixed/near-normal temperature pattern.")
    summary = " ".join(parts)
    if "bearish" in bias:
        reason = "Gas-weighted temperature signal leans cooler in key demand regions. In shoulder/injection season this can suppress early cooling demand."
    elif "bullish" in bias:
        reason = "Gas-weighted temperature signal leans warmer in key demand regions. If persistent, this can raise power burn and gas demand."
    else:
        reason = "Temperature pattern is mixed across gas demand regions. No dominant weather-driven demand signal from this map alone."
    return {"score": final_score, "bias": bias, "summary": summary, "reason": reason, "region_scores": region_scores, "composite_raw": round(composite, 3), "analyzer": "v0.5 noaa color heuristic", "generated_at": datetime.now().isoformat()}


def analyze_chart_image_basic(image_path: Path, timeframe: str) -> dict:
    img = Image.open(image_path).convert("RGB")
    arr = np.array(img).astype(np.float32)
    h, w, _ = arr.shape
    crop = arr[int(h * 0.12):int(h * 0.90), int(w * 0.08):int(w * 0.92)]
    ch, cw, _ = crop.shape
    r, g, b = crop[:, :, 0], crop[:, :, 1], crop[:, :, 2]
    green = (g > r + 18) & (g > b + 5) & (g > 90)
    red = (r > g + 18) & (r > b + 5) & (r > 90)
    mask = green | red
    recent = mask[:, int(cw * 0.65):]
    recent_green = green[:, int(cw * 0.65):].sum()
    recent_red = red[:, int(cw * 0.65):].sum()
    total_recent = max(recent.sum(), 1)
    red_ratio, green_ratio = recent_red / total_recent, recent_green / total_recent
    ys, xs = np.where(recent)
    mean_y = ys.mean() / ch if len(ys) > 20 else 0.50
    left = mask[:, int(cw * 0.25):int(cw * 0.55)]
    ly, lx = np.where(left)
    slope = (ly.mean() / ch - mean_y) if len(ly) > 20 and len(ys) > 20 else 0.0
    raw = (green_ratio - red_ratio) * 1.2 + slope * 2.0
    if raw >= 0.45:
        score, bias, structure = 1, "slightly bullish", "recent structure leans upward / buying pressure visible"
    elif raw <= -0.45:
        score, bias, structure = -1, "slightly bearish", "recent structure leans downward / selling pressure visible"
    else:
        score, bias, structure = 0, "neutral", "mixed or range-like structure"
    if mean_y > 0.68 and red_ratio > 0.55:
        score, bias, structure = -2, "bearish", "recent candles are low in the chart with red dominance; weak/breakdown structure"
    elif mean_y < 0.32 and green_ratio > 0.55:
        score, bias, structure = 2, "bullish", "recent candles are high in the chart with green dominance; strong/breakout structure"
    return {"timeframe": timeframe, "score": score, "bias": bias, "summary": f"{timeframe}: {structure}.", "reason": f"Auto chart heuristic: green={green_ratio:.2f}, red={red_ratio:.2f}, placement={mean_y:.2f}, slope={slope:.2f}. Verify manually against ZigZag HH/HL/LH/LL, VWAP, and volume.", "metrics": {"green_ratio": round(float(green_ratio),3), "red_ratio": round(float(red_ratio),3), "recent_vertical_placement": round(float(mean_y),3), "rough_slope": round(float(slope),3), "raw": round(float(raw),3)}, "analyzer": "v0.5 chart heuristic", "generated_at": datetime.now().isoformat()}


def build_cme_composite(mod: dict) -> dict:
    """
    v0.6 regime-first CME structure engine.
    Higher timeframe dominates. Lower timeframe bullishness inside a bearish 1D regime is treated as pullback,
    not reversal, unless 1D itself turns bullish.
    """
    ensure_cme_structure(mod)

    weights = {"1H": 0.20, "4H": 0.30, "1D": 0.50}
    weighted = 0.0
    used_weight = 0.0
    parts = []
    tf_scores = {}

    for tf, wt in weights.items():
        tfdata = mod["cme_timeframes"][tf]
        if tfdata.get("image_path"):
            sc = int(tfdata.get("score", 0))
            tf_scores[tf] = sc
            weighted += sc * wt
            used_weight += wt
            if tfdata.get("summary"):
                parts.append(tfdata["summary"])
        else:
            tf_scores[tf] = 0

    raw = weighted / used_weight if used_weight else 0.0

    def label(score_value: float) -> str:
        if score_value >= 1.4:
            return "bullish"
        if score_value >= 0.5:
            return "slightly bullish"
        if score_value <= -1.4:
            return "bearish"
        if score_value <= -0.5:
            return "slightly bearish"
        return "neutral"

    d1 = tf_scores.get("1D", 0)
    h4 = tf_scores.get("4H", 0)
    h1 = tf_scores.get("1H", 0)

    regime_note = "No dominant 1D regime. Composite uses weighted 1H/4H/1D."
    final_raw = raw

    # HTF override: bearish 1D dominates; bullish 1H/4H becomes pullback unless 1D flips.
    if d1 <= -1:
        final_raw = min(raw, -0.75)
        regime_note = "1D bearish regime: lower-timeframe bullish signals are treated as pullbacks, not reversal."
    elif d1 >= 1:
        final_raw = max(raw, 0.75)
        regime_note = "1D bullish regime: lower-timeframe bearish signals are treated as pullbacks, not reversal."
    else:
        # If 1D is neutral but 4H and 1H agree, allow tactical bias.
        if h4 <= -1 and h1 <= -1:
            final_raw = min(raw, -0.75)
            regime_note = "1D neutral, but 4H and 1H align bearish: tactical bearish structure."
        elif h4 >= 1 and h1 >= 1:
            final_raw = max(raw, 0.75)
            regime_note = "1D neutral, but 4H and 1H align bullish: tactical bullish structure."

    final_bias = label(final_raw)
    if final_bias == "bullish":
        final_score = 2
    elif final_bias == "slightly bullish":
        final_score = 1
    elif final_bias == "bearish":
        final_score = -2
    elif final_bias == "slightly bearish":
        final_score = -1
    else:
        final_score = 0

    summary = " ".join(parts) if parts else "No CME timeframe images analyzed yet."
    reason = (
        f"Regime-first CME score={final_raw:.2f}. Raw weighted score={raw:.2f}. "
        f"Weights: 1D 50%, 4H 30%, 1H 20%. "
        f"TF scores: 1D={d1}, 4H={h4}, 1H={h1}. {regime_note}"
    )

    return {
        "score": final_score,
        "bias": final_bias,
        "summary": summary,
        "reason": reason,
        "composite_raw": round(raw, 3),
        "regime_first_score": round(final_raw, 3),
        "tf_scores": tf_scores,
        "engine": "v0.6 regime-first CME structure",
    }


def score_text_for_gas(text: str) -> dict:
    t = text.lower()
    bearish, bullish = 0, 0
    bear_hits, bull_hits = [], []
    for term, wt in BEARISH_TERMS.items():
        if term in t:
            bearish += wt
            bear_hits.append(term)
    for term, wt in BULLISH_TERMS.items():
        if term in t:
            bullish += wt
            bull_hits.append(term)
    net = bullish - bearish
    if net >= 5: score, bias = 2, "bullish"
    elif net >= 2: score, bias = 1, "slightly bullish"
    elif net <= -5: score, bias = -2, "bearish"
    elif net <= -2: score, bias = -1, "slightly bearish"
    else: score, bias = 0, "neutral"
    return {"score": score, "bias": bias, "bullish_points": bullish, "bearish_points": bearish, "bullish_hits": bull_hits, "bearish_hits": bear_hits, "net": net}


def fetch_url_text(url: str, timeout: int = 10) -> dict:
    headers = {"User-Agent": "Mozilla/5.0 XNGResearchDesk/0.5"}
    r = requests.get(url, headers=headers, timeout=timeout)
    soup = BeautifulSoup(r.text, "html.parser")
    title = (soup.title.string.strip() if soup.title and soup.title.string else "")
    meta = ""
    tag = soup.find("meta", attrs={"name": "description"}) or soup.find("meta", attrs={"property": "og:description"})
    if tag and tag.get("content"):
        meta = tag["content"].strip()
    paragraphs = " ".join(p.get_text(" ", strip=True) for p in soup.find_all("p")[:12])
    return {"url": url, "title": title, "description": meta, "text": paragraphs[:5000]}


def fetch_google_news(query: str, limit: int = 8) -> list:
    rss = f"https://news.google.com/rss/search?q={quote_plus(query)}&hl=en-US&gl=US&ceid=US:en"
    items = []

    if feedparser is not None:
        feed = feedparser.parse(rss)
        for e in feed.entries[:limit]:
            items.append({
                "title": getattr(e, "title", ""),
                "url": getattr(e, "link", ""),
                "source": getattr(getattr(e, "source", None), "title", ""),
                "published": getattr(e, "published", ""),
            })
        return items

    # fallback: no feedparser required
    import xml.etree.ElementTree as ET
    headers = {"User-Agent": "Mozilla/5.0 XNGResearchDesk/0.5.1"}
    r = requests.get(rss, headers=headers, timeout=12)
    r.raise_for_status()
    root = ET.fromstring(r.content)
    channel = root.find("channel")
    if channel is None:
        return items
    for item in channel.findall("item")[:limit]:
        title = item.findtext("title") or ""
        link = item.findtext("link") or ""
        pub = item.findtext("pubDate") or ""
        source_el = item.find("{*}source")
        source = source_el.text if source_el is not None and source_el.text else ""
        items.append({"title": title, "url": link, "source": source, "published": pub})
    return items


def analyze_news_items(items: list) -> dict:
    analyzed = []
    total = 0

    for i, item in enumerate(items, start=1):
        text = " ".join([
            item.get("title", ""),
            item.get("description", ""),
            item.get("text", ""),
        ])
        sc = score_text_for_gas(text)
        analyzed.append({**item, **sc, "index": i})
        total += sc["score"]

    avg = total / max(len(analyzed), 1)

    if avg >= 1.4:
        score, bias = 2, "bullish"
    elif avg >= 0.75:
        score, bias = 1, "slightly bullish"
    elif avg <= -1.4:
        score, bias = -2, "bearish"
    elif avg <= -0.75:
        score, bias = -1, "slightly bearish"
    else:
        score, bias = 0, "neutral"

    bearish_items = [x for x in analyzed if x["score"] < 0]
    bullish_items = [x for x in analyzed if x["score"] > 0]
    neutral_items = [x for x in analyzed if x["score"] == 0]

    summary = (
        f"News composite: {bias}. "
        f"Analyzed {len(analyzed)} items: "
        f"{len(bearish_items)} bearish, {len(bullish_items)} bullish, {len(neutral_items)} neutral."
    )

    item_lines = []
    for x in analyzed:
        item_lines.append(
            f"{x['index']}. [{x['bias']} / score {x['score']}] {x.get('title','')} "
            f"(bull={x.get('bullish_points',0)}, bear={x.get('bearish_points',0)})"
        )

    reason = (
        f"Average item score={avg:.2f}. "
        "Item-level classification:\n" + "\n".join(item_lines)
    )

    return {
        "score": score,
        "bias": bias,
        "summary": summary,
        "reason": reason,
        "items": analyzed,
        "avg_score": round(avg, 3),
        "counts": {
            "total": len(analyzed),
            "bearish": len(bearish_items),
            "bullish": len(bullish_items),
            "neutral": len(neutral_items),
        },
        "generated_at": datetime.now().isoformat(),
    }



def markdown_to_txt(md: str) -> str:
    txt = md
    txt = re.sub(r"^#{1,6}\s*", "", txt, flags=re.MULTILINE)
    txt = txt.replace("**", "").replace("*", "")
    txt = txt.replace("✅", "[DONE]").replace("⬜", "[TODO]")
    txt = re.sub(r"\n{3,}", "\n\n", txt)
    return txt.strip() + "\n"

def build_report(data: dict) -> str:
    lines = [f"# XNG Daily Data Report — {data['date']}", "", "## Completion"]
    for m in MODULES:
        status = "✅" if data["modules"][m]["completed"] else "⬜"
        lines.append(f"- {status} {m}")
    lines += ["", "## Module Outputs"]
    for m in MODULES:
        mod = data["modules"][m]
        rv = report_values(mod)
        display_score = rv["score"]
        display_bias = rv["bias"]
        display_summary = rv["summary"]
        display_reason = rv["reason"]
        ar = rv["auto_result"]

        lines += [f"### {m}", f"- Completed: {mod['completed']}", f"- Score: {display_score}", f"- Bias: {display_bias}"]
        if display_summary: lines.append(f"- Summary: {display_summary}")
        if display_reason: lines.append(f"- Reason: {display_reason}")
        if mod.get("raw_notes"): lines.append(f"- Notes: {mod['raw_notes']}")

        if m == "News Scan" and ar.get("items"):
            lines.append("")
            lines.append("#### News Items")
            for item in ar["items"]:
                title = item.get("title", "")
                source = item.get("source", "")
                item_bias = item.get("bias", "")
                item_score = item.get("score", "")
                bull = item.get("bullish_points", 0)
                bear = item.get("bearish_points", 0)
                lines.append(f"- {item.get('index','')}. [{item_bias} / {item_score}] {title} | {source} | bull={bull}, bear={bear}")
        lines.append("")
    ensure_weekly_storage(data)
    weekly = data["weekly"][WEEKLY_MODULE]

    lines += ["## Weekly Storage"]
    lines += [
        f"- Completed: {weekly.get('completed', False)}",
        f"- Score: {weekly.get('score', 0)}",
        f"- Bias: {weekly.get('bias', 'neutral')}",
        f"- Week ending: {weekly.get('week_ending', '')}",
    ]
    if weekly.get("summary"):
        lines.append(f"- Summary: {weekly['summary']}")
    if weekly.get("reason"):
        lines.append(f"- Reason: {weekly['reason']}")
    if weekly.get("raw_notes"):
        lines.append(f"- Notes: {weekly['raw_notes']}")
    lines.append("")

    ensure_monthly_steo(data)
    monthly = data["monthly"][MONTHLY_MODULE]

    lines += ["## Monthly STEO"]
    lines += [
        f"- Completed: {monthly.get('completed', False)}",
        f"- Score: {monthly.get('score', 0)}",
        f"- Bias: {monthly.get('bias', 'neutral')}",
        f"- Release month/date: {monthly.get('release_month', '')}",
    ]
    if monthly.get("summary"):
        lines.append(f"- Summary: {monthly['summary']}")
    if monthly.get("reason"):
        lines.append(f"- Reason: {monthly['reason']}")
    if monthly.get("raw_notes"):
        lines.append(f"- Notes: {monthly['raw_notes']}")
    lines.append("")

    noaa_scores = [
        int(report_values(data["modules"]["NOAA 6-10D"])["score"]),
        int(report_values(data["modules"]["NOAA 8-14D"])["score"]),
    ]
    noaa_score = sum(noaa_scores) / 2
    cme_score = int(report_values(data["modules"]["CME NG Structure"])["score"])
    news_score = int(report_values(data["modules"]["News Scan"])["score"])
    weekly_score = int(weekly.get("score", 0))
    monthly_score = int(monthly.get("score", 0))

    # v0.8: Monthly STEO is a regime background layer.
    # It is intentionally capped at 15% so it guides context but does not dominate weekly/daily event signals.
    composite = (
        noaa_score * 0.30
        + cme_score * 0.25
        + news_score * 0.10
        + weekly_score * 0.20
        + monthly_score * 0.15
    )

    final_bias = final_bias_from_score(composite)
    confidence = confidence_from_score(composite)
    lines += [
        "## Draft Daily Bias",
        f"- Composite: {composite:.2f}",
        f"- Auto final bias: {final_bias}",
        f"- Confidence: {confidence}",
        f"- NOAA component: {noaa_score:.2f} × 30%",
        f"- CME component: {cme_score} × 25%",
        f"- News component: {news_score} × 10%",
        f"- Weekly Storage component: {weekly_score} × 20%",
        f"- Monthly STEO component: {monthly_score} × 15%",
        "- Regime note: Monthly STEO is background context, not an execution trigger.",
        "- Final bias: auto-calculated; review manually before execution.",
        "",
    ]
    return "\n".join(lines)



def ocr_image_text(image_path: Path) -> str:
    """
    OCR helper for EIA screenshots.
    Requires pytesseract + system tesseract installed on the user's Mac.
    If unavailable, user can paste Summary Text into the fallback text area.
    """
    try:
        import pytesseract
    except Exception as e:
        raise RuntimeError("pytesseract is not installed. Use Paste EIA text fallback, or install pytesseract + tesseract.") from e

    img = Image.open(image_path).convert("RGB")
    return pytesseract.image_to_string(img)


def parse_eia_storage_text(text: str) -> dict:
    """
    Parse EIA Weekly Natural Gas Storage report text.
    Works with official summary wording and many OCR outputs.
    """
    clean = " ".join(text.replace(",", "").split())

    total_storage = None
    actual = None
    last_year = None
    five_year_avg = None

    # Example:
    # Working gas in storage was 2063 Bcf ... net increase of 103 Bcf ...
    m = re.search(r"storage was\s+([0-9]+(?:\.[0-9]+)?)\s+Bcf", clean, flags=re.I)
    if m:
        total_storage = float(m.group(1))

    m = re.search(r"net (?:increase|decrease) of\s+([0-9]+(?:\.[0-9]+)?)\s+Bcf", clean, flags=re.I)
    if m:
        actual = float(m.group(1))
        if re.search(r"net decrease of", clean, flags=re.I):
            actual = -actual

    # Example:
    # Stocks were 142 Bcf higher than last year ... and 137 Bcf above the five-year average of 1926 Bcf.
    m = re.search(r"Stocks were\s+([0-9]+(?:\.[0-9]+)?)\s+Bcf\s+(higher|lower)\s+than last year", clean, flags=re.I)
    if m and total_storage is not None:
        diff = float(m.group(1))
        last_year = total_storage - diff if m.group(2).lower() == "higher" else total_storage + diff

    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s+Bcf\s+(above|below)\s+the five-year average(?: of\s+([0-9]+(?:\.[0-9]+)?))?", clean, flags=re.I)
    if m and total_storage is not None:
        diff = float(m.group(1))
        if m.group(3):
            five_year_avg = float(m.group(3))
        else:
            five_year_avg = total_storage - diff if m.group(2).lower() == "above" else total_storage + diff

    # Table fallback: Total 2063 1960 103 103 1921 7.4 1926 7.1
    if any(v is None for v in [total_storage, actual, last_year, five_year_avg]):
        m = re.search(r"Total\s+([0-9]+)\s+([0-9]+)\s+R?\s+([0-9\-]+)\s+([0-9\-]+)\s+([0-9]+)\s+[-0-9.]+\s+([0-9]+)\s+[-0-9.]+", clean, flags=re.I)
        if m:
            total_storage = total_storage if total_storage is not None else float(m.group(1))
            actual = actual if actual is not None else float(m.group(3))
            last_year = last_year if last_year is not None else float(m.group(5))
            five_year_avg = five_year_avg if five_year_avg is not None else float(m.group(6))

    missing = []
    if total_storage is None: missing.append("total_storage")
    if actual is None: missing.append("actual_net_change")
    if last_year is None: missing.append("last_year")
    if five_year_avg is None: missing.append("five_year_avg")

    return {
        "ok": len(missing) == 0,
        "missing": missing,
        "actual": actual or 0.0,
        "total_storage": total_storage or 0.0,
        "last_year": last_year or 0.0,
        "five_year_avg": five_year_avg or 0.0,
        "raw_text_preview": clean[:1200],
    }



def fetch_latest_eia_storage_report() -> dict:
    """
    Fetch latest EIA Weekly Natural Gas Storage Report directly from EIA HTML.
    No OCR required.
    """
    url = "https://ir.eia.gov/ngs/ngs.html"
    headers = {"User-Agent": "Mozilla/5.0 XNGResearchDesk/0.7.2"}
    r = requests.get(url, headers=headers, timeout=15)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")
    page_text = soup.get_text(" ", strip=True)

    parsed = parse_eia_storage_text(page_text)

    title = ""
    h1 = soup.find(["h1", "h2"])
    if h1:
        title = h1.get_text(" ", strip=True)

    release = ""
    m = re.search(r"Released:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4}\s+at\s+[0-9:]+\s+[ap]\.?m\.?)", page_text, flags=re.I)
    if m:
        release = m.group(1)

    week_ending = ""
    m = re.search(r"for week ending\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})", page_text, flags=re.I)
    if m:
        week_ending = m.group(1)
    else:
        m = re.search(r"week ending\s+([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", page_text, flags=re.I)
        if m:
            week_ending = m.group(1)

    return {
        "url": url,
        "title": title or "Weekly Natural Gas Storage Report",
        "release": release,
        "week_ending": week_ending,
        "page_text": page_text[:12000],
        "parsed": parsed,
        "fetched_at": datetime.now().isoformat(),
    }


def analyze_eia_html_fetch(expected: float) -> dict:
    fetched = fetch_latest_eia_storage_report()
    parsed = fetched["parsed"]

    result = analyze_weekly_storage(
        actual=float(parsed["actual"]),
        expected=float(expected),
        total_storage=float(parsed["total_storage"]),
        five_year_avg=float(parsed["five_year_avg"]),
        last_year=float(parsed["last_year"]),
    )

    result["source"] = "eia_html_fetch"
    result["url"] = fetched["url"]
    result["title"] = fetched["title"]
    result["release"] = fetched["release"]
    result["week_ending"] = fetched.get("week_ending", "")
    result["parse"] = parsed
    result["extracted_text"] = fetched["page_text"][:5000]

    if not parsed["ok"]:
        result["reason"] += " Parser warning: missing fields = " + ", ".join(parsed["missing"]) + ". Review manually."

    return result


def analyze_eia_image_or_text(image_path: Path | None, pasted_text: str, expected: float) -> dict:
    if pasted_text.strip():
        text = pasted_text
        source = "pasted_text"
    elif image_path is not None:
        text = ocr_image_text(image_path)
        source = "ocr_image"
    else:
        raise RuntimeError("Upload EIA screenshot or paste EIA summary text first.")

    parsed = parse_eia_storage_text(text)
    result = analyze_weekly_storage(
        actual=float(parsed["actual"]),
        expected=float(expected),
        total_storage=float(parsed["total_storage"]),
        five_year_avg=float(parsed["five_year_avg"]),
        last_year=float(parsed["last_year"]),
    )

    result["parse"] = parsed
    result["source"] = source
    result["extracted_text"] = text[:5000]
    if not parsed["ok"]:
        result["reason"] += " Parser warning: missing fields = " + ", ".join(parsed["missing"]) + ". Review manually."
    return result


st.set_page_config(page_title="XNG Research Desk", layout="wide")
st.title("XNG Research Desk v0.9.4 — NOAA Shapefile Primary")
st.caption("NOAA 6-10D/8-14D now use NOAA CPC shapefile primary engine. CME + News + Weekly Storage + Monthly STEO unchanged.")

selected_day = st.date_input("Research date", value=date.today()).isoformat()
data = load_day(selected_day)

with st.sidebar:
    st.header("Daily Modules")
    for m in MODULES:
        st.write(("✅" if data["modules"][m]["completed"] else "⬜") + f" {m}")
    st.divider()
    if st.button("Generate / Refresh Daily Report", width="stretch"):
        data["daily_report"] = build_report(data); save_day(selected_day, data); st.success("Daily report generated.")
    if st.button("Reset Today Data", width="stretch"):
        fp = day_file(selected_day)
        if fp.exists(): fp.unlink()
        data = make_new_day(selected_day); save_day(selected_day, data); st.success("Today data reset."); st.rerun()

tabs = st.tabs(MODULES + [WEEKLY_MODULE, MONTHLY_MODULE, "Daily Report Center"])

for idx, module_name in enumerate(MODULES):
    with tabs[idx]:
        mod = data["modules"][module_name]
        st.subheader(module_name)
        left, right = st.columns([1,1])

        with left:
            if module_name == "CME NG Structure":
                ensure_cme_structure(mod)
                for tf in ["1H","4H","1D"]:
                    st.markdown(f"#### {tf}")
                    up = st.file_uploader(f"Drop {tf} chart screenshot", type=["png","jpg","jpeg","webp"], key=f"upload_cme_{tf}")
                    if up is not None:
                        mod["cme_timeframes"][tf]["image_path"] = save_uploaded_image(selected_day, f"CME NG Structure {tf}", up)
                        save_day(selected_day, data); st.success(f"Saved {tf}")
                    rel = mod["cme_timeframes"][tf].get("image_path")
                    if rel and (ROOT/rel).exists():
                        st.image(str(ROOT/rel), caption=f"{tf}: {rel}", width="stretch")
            elif module_name == "News Scan":
                ensure_news(mod)
                up = st.file_uploader("Drop news screenshot for archive only", type=["png","jpg","jpeg","webp"], key="upload_news")
                if up is not None:
                    mod["image_path"] = save_uploaded_image(selected_day, module_name, up)
                    save_day(selected_day, data); st.success("Saved news screenshot")
                if mod.get("image_path") and (ROOT/mod["image_path"]).exists():
                    st.image(str(ROOT/mod["image_path"]), caption=mod["image_path"], width="stretch")
            else:
                if module_name.startswith("NOAA"):
                    st.markdown("#### NOAA CPC Shapefile Primary")
                    st.caption("Upload NOAA shapefile ZIP (recommended on iPhone) or select .shp + .dbf + .shx together. This replaces screenshot OCR/warp.")
                    shp_files = st.file_uploader(
                        f"Drop NOAA shapefile files for {module_name}",
                        type=["zip", "shp", "dbf", "shx", "prj", "cpg", "sbn", "sbx", "xml"],
                        accept_multiple_files=True,
                        key=f"upload_shp_{slugify(module_name)}",
                    )
                    if shp_files:
                        try:
                            saved = save_uploaded_noaa_shapefile_set(selected_day, module_name, shp_files)
                            mod["shapefile_set"] = saved
                            mod["source_path"] = saved.get("shp_path")
                            save_day(selected_day, data)
                            st.success(f"Saved NOAA shapefile: {saved.get('shp_path')}")
                        except Exception as e:
                            st.error(f"NOAA shapefile save failed: {e}")

                    if mod.get("source_path"):
                        st.info(f"NOAA shapefile source: {mod['source_path']}")
                        with st.expander("Uploaded shapefile files"):
                            st.json(mod.get("shapefile_set", {}))

                    st.divider()
                    st.markdown("#### Legacy screenshot archive only")
                    up = st.file_uploader(f"Optional screenshot archive for {module_name}", type=["png","jpg","jpeg","webp"], key=f"upload_img_{slugify(module_name)}")
                    if up is not None:
                        mod["image_path"] = save_uploaded_image(selected_day, module_name, up)
                        save_day(selected_day, data); st.success(f"Saved screenshot archive: {mod['image_path']}")
                    if mod.get("image_path") and (ROOT/mod["image_path"]).exists():
                        st.image(str(ROOT/mod["image_path"]), caption=mod["image_path"], width="stretch")
                else:
                    up = st.file_uploader(f"Drop screenshot for {module_name}", type=["png","jpg","jpeg","webp"], key=f"upload_{slugify(module_name)}")
                    if up is not None:
                        mod["image_path"] = save_uploaded_image(selected_day, module_name, up)
                        save_day(selected_day, data); st.success(f"Saved image: {mod['image_path']}")
                    if mod.get("image_path") and (ROOT/mod["image_path"]).exists():
                        st.image(str(ROOT/mod["image_path"]), caption=mod["image_path"], width="stretch")

        with right:
            st.markdown("### Interpretation")

            if module_name.startswith("NOAA"):
                if st.button("Auto Fill from NOAA Shapefile", key=f"auto_{slugify(module_name)}", width="stretch"):
                    if not mod.get("source_path"):
                        st.warning("Upload the NOAA shapefile ZIP or the .shp/.dbf/.shx set first.")
                    else:
                        try:
                            res = analyze_noaa_temperature_shapefile(ROOT/mod["source_path"], module_name)
                            st.session_state[f"preview_{slugify(module_name)}"] = res
                            st.success("NOAA shapefile preview generated.")
                        except Exception as e:
                            st.error(f"NOAA shapefile analysis failed: {e}")
                pk = f"preview_{slugify(module_name)}"
                if pk in st.session_state:
                    st.json(st.session_state[pk])
                    if st.button("Apply NOAA Shapefile to Fields", key=f"apply_{slugify(module_name)}", width="stretch"):
                        res = st.session_state[pk]
                        mod.update({k: res[k] for k in ["score","bias","summary","reason"]})
                        mod["auto_result"] = res
                        mod["field_version"] = int(mod.get("field_version", 0)) + 1
                        save_day(selected_day, data)
                        del st.session_state[pk]
                        st.rerun()

            if module_name == "CME NG Structure":
                ensure_cme_structure(mod)
                for tf in ["1H","4H","1D"]:
                    if st.button(f"Auto Extract {tf}", key=f"auto_cme_{tf}", width="stretch"):
                        rel = mod["cme_timeframes"][tf].get("image_path")
                        if not rel: st.warning(f"Upload {tf} first.")
                        else:
                            res = analyze_chart_image_basic(ROOT/rel, tf)
                            st.session_state[f"preview_cme_{tf}"] = res; st.success(f"{tf} preview generated.")
                    pk = f"preview_cme_{tf}"
                    if pk in st.session_state:
                        st.json(st.session_state[pk])
                        if st.button(f"Apply {tf} Preview", key=f"apply_cme_{tf}", width="stretch"):
                            res = st.session_state[pk]
                            mod["cme_timeframes"][tf].update({k: res[k] for k in ["score","bias","summary","reason"]}); mod["cme_timeframes"][tf]["auto_result"] = res
                            save_day(selected_day, data); del st.session_state[pk]; st.rerun()
                if st.button("Build CME Composite Output", key="build_cme_composite", width="stretch"):
                    comp = build_cme_composite(mod)
                    mod.update({k: comp[k] for k in ["score","bias","summary","reason"]}); mod["auto_result"] = {"composite": comp, "timeframes": mod.get("cme_timeframes", {})}
                    mod["field_version"] = int(mod.get("field_version", 0)) + 1
                    save_day(selected_day, data); st.success("Built CME composite."); st.rerun()

            if module_name == "News Scan":
                ensure_news(mod)
                mod["news_query"] = st.text_input("Google News query", value=mod.get("news_query","US natural gas news"))
                if st.button("Fetch Google News RSS", width="stretch"):
                    try:
                        items = fetch_google_news(mod["news_query"], limit=10)
                        mod["news_items"] = items
                        save_day(selected_day, data); st.success(f"Fetched {len(items)} news items.")
                    except Exception as e:
                        st.error(f"Fetch failed: {e}")
                manual_urls = st.text_area("Manual URLs, one per line", height=90)
                if st.button("Read Manual URLs", width="stretch"):
                    items = []
                    for url in [u.strip() for u in manual_urls.splitlines() if u.strip()]:
                        try:
                            items.append(fetch_url_text(url))
                        except Exception as e:
                            items.append({"url": url, "title": "FETCH FAILED", "description": str(e), "text": ""})
                    mod["news_items"] = mod.get("news_items", []) + items
                    save_day(selected_day, data); st.success(f"Read {len(items)} manual URLs.")
                if mod.get("news_items"):
                    st.markdown("#### News items")
                    st.json(mod["news_items"][:10])
                    if st.button("Analyze News Items", width="stretch"):
                        res = analyze_news_items(mod["news_items"])
                        mod.update({k: res[k] for k in ["score","bias","summary","reason"]}); mod["auto_result"] = res
                        mod["field_version"] = int(mod.get("field_version", 0)) + 1
                        save_day(selected_day, data); st.success("News analyzed."); st.rerun()

            st.divider()
            field_version = int(mod.get("field_version", 0))
            field_key = f"{slugify(module_name)}_{field_version}"
            manual_mode = st.toggle("Manual fill / edit mode", value=True, key=f"manual_{field_key}")
            disabled = not manual_mode
            st.caption(f"Field version: {field_version}")
            score_value = st.slider("Score (-2 bearish to +2 bullish)", -2, 2, int(mod.get("score", 0)), key=f"score_{field_key}", disabled=disabled)
            bias_value = mod.get("bias","neutral") if mod.get("bias","neutral") in BIAS_OPTIONS else "neutral"
            selected_bias = st.selectbox("Bias", BIAS_OPTIONS, index=BIAS_OPTIONS.index(bias_value), key=f"bias_{field_key}", disabled=disabled)
            summary_value = st.text_area("Output summary", value=mod.get("summary",""), height=110, key=f"summary_{field_key}", disabled=disabled)
            reason_value = st.text_area("Reason / interpretation", value=mod.get("reason",""), height=150, key=f"reason_{field_key}", disabled=disabled)
            notes_value = st.text_area("Raw notes", value=mod.get("raw_notes",""), height=100, key=f"notes_{field_key}", disabled=disabled)
            a,b = st.columns(2)
            if a.button("Save Draft", key=f"save_{slugify(module_name)}", width="stretch"):
                mod.update(choose_final_values(mod, score_value, selected_bias, summary_value, reason_value, notes_value))
                save_day(selected_day, data); st.success("Saved draft.")
            if b.button("Confirm Complete ✅", key=f"confirm_{slugify(module_name)}", width="stretch"):
                mod.update(choose_final_values(mod, score_value, selected_bias, summary_value, reason_value, notes_value))
                mod.update({"completed": True, "confirmed_at": datetime.now().isoformat()})
                save_day(selected_day, data); st.success(f"{module_name} completed."); st.rerun()

        if mod.get("completed"):
            st.success(f"{module_name} is completed.")
        else:
            st.warning(f"{module_name} is not completed yet.")


with tabs[-3]:
    ensure_weekly_storage(data)
    weekly = data["weekly"][WEEKLY_MODULE]

    st.subheader("Weekly Storage")
    st.caption("Upload EIA Weekly Natural Gas Storage Report screenshot, extract values, review, then apply.")

    left, right = st.columns([1, 1])

    with left:
        st.markdown("### EIA report image")
        up = st.file_uploader(
            "Drop EIA Weekly Storage screenshot",
            type=["png", "jpg", "jpeg", "webp"],
            key="upload_weekly_eia",
        )
        if up is not None:
            weekly["image_path"] = save_uploaded_image(selected_day, "Weekly Storage EIA", up)
            save_day(selected_day, data)
            st.success(f"Saved EIA image: {weekly['image_path']}")

        if weekly.get("image_path"):
            img_path = ROOT / weekly["image_path"]
            if img_path.exists():
                st.image(str(img_path), caption=weekly["image_path"], width="stretch")

        st.markdown("### OCR fallback")
        st.caption("If OCR is unavailable, paste the EIA Summary text here. The parser will use pasted text first.")
        pasted_text = st.text_area(
            "Paste EIA summary text / table text",
            value=weekly.get("extracted_text", ""),
            height=180,
            key=f"weekly_paste_{weekly.get('field_version',0)}",
        )

    with right:
        st.markdown("### Weekly Storage Inputs")

        c1, c2 = st.columns(2)
        with c1:
            weekly["week_ending"] = st.text_input("Week ending", value=weekly.get("week_ending", ""))
            st.caption("Auto-filled from EIA fetch when available. Optional metadata; not used in score.")
            expected = st.number_input("Expected / consensus (Bcf)", value=float(weekly.get("expected", 0.0)), step=1.0)

        with c2:
            st.metric("Current weekly score", weekly.get("score", 0))
            st.metric("Current weekly bias", weekly.get("bias", "neutral"))

        if st.button("Fetch Latest EIA Report (no OCR)", width="stretch"):
            try:
                res = analyze_eia_html_fetch(expected)
                st.session_state["weekly_preview"] = res
                st.success("Latest EIA report fetched and preview generated.")
            except Exception as e:
                st.error(f"EIA fetch failed: {e}")

        if st.button("Auto Extract Weekly Storage from Image/Text", width="stretch"):
            try:
                image_path = ROOT / weekly["image_path"] if weekly.get("image_path") else None
                res = analyze_eia_image_or_text(image_path, pasted_text, expected)
                st.session_state["weekly_preview"] = res
                st.success("Weekly extraction preview generated.")
            except Exception as e:
                st.error(f"Weekly extraction failed: {e}")

        if "weekly_preview" in st.session_state:
            st.json(st.session_state["weekly_preview"])
            if st.button("Apply Weekly Extraction to Fields", width="stretch"):
                res = st.session_state["weekly_preview"]
                metrics = res["metrics"]
                weekly.update({
                    "week_ending": weekly.get("week_ending") or res.get("week_ending", ""),
                    "expected": expected,
                    "actual": metrics["actual"],
                    "total_storage": metrics["total_storage"],
                    "five_year_avg": metrics["five_year_avg"],
                    "last_year": metrics["last_year"],
                    "surprise": metrics["surprise"],
                    "vs_5y": metrics["vs_5y"],
                    "vs_last_year": metrics["vs_last_year"],
                    "score": res["score"],
                    "bias": res["bias"],
                    "summary": res["summary"],
                    "reason": res["reason"],
                    "extracted_text": res.get("extracted_text", ""),
                    "auto_result": res,
                    "field_version": int(weekly.get("field_version", 0)) + 1,
                })
                save_day(selected_day, data)
                del st.session_state["weekly_preview"]
                st.success("Weekly extraction applied.")
                st.rerun()

        st.divider()
        st.markdown("### Manual weekly review / override")
        fv = int(weekly.get("field_version", 0))
        manual_weekly = st.toggle("Manual weekly edit mode", value=True, key=f"manual_weekly_{fv}")
        disabled_weekly = not manual_weekly

        c3, c4 = st.columns(2)
        with c3:
            actual = st.number_input("Actual net change (Bcf)", value=float(weekly.get("actual", 0.0)), step=1.0, disabled=disabled_weekly)
            total_storage = st.number_input("Total storage (Bcf)", value=float(weekly.get("total_storage", 0.0)), step=1.0, disabled=disabled_weekly)
            five_year_avg = st.number_input("5-year average stocks (Bcf)", value=float(weekly.get("five_year_avg", 0.0)), step=1.0, disabled=disabled_weekly)
            last_year = st.number_input("Year ago stocks (Bcf)", value=float(weekly.get("last_year", 0.0)), step=1.0, disabled=disabled_weekly)
        with c4:
            score_value = st.slider("Weekly Score (-2 bearish to +2 bullish)", -2, 2, int(weekly.get("score", 0)), key=f"weekly_score_{fv}", disabled=disabled_weekly)
            bias_value = weekly.get("bias", "neutral") if weekly.get("bias", "neutral") in BIAS_OPTIONS else "neutral"
            selected_bias = st.selectbox("Weekly Bias", BIAS_OPTIONS, index=BIAS_OPTIONS.index(bias_value), key=f"weekly_bias_{fv}", disabled=disabled_weekly)

        summary_value = st.text_area("Weekly summary", value=weekly.get("summary", ""), height=90, key=f"weekly_summary_{fv}", disabled=disabled_weekly)
        reason_value = st.text_area("Weekly reason", value=weekly.get("reason", ""), height=110, key=f"weekly_reason_{fv}", disabled=disabled_weekly)
        notes_value = st.text_area("Weekly raw notes", value=weekly.get("raw_notes", ""), height=80, key=f"weekly_notes_{fv}", disabled=disabled_weekly)

        a, b = st.columns(2)
        if a.button("Save Weekly Draft", width="stretch"):
            weekly.update({
                "actual": actual,
                "expected": expected,
                "total_storage": total_storage,
                "five_year_avg": five_year_avg,
                "last_year": last_year,
                "surprise": actual - expected,
                "vs_5y": total_storage - five_year_avg,
                "vs_last_year": total_storage - last_year,
                "score": score_value,
                "bias": selected_bias,
                "summary": summary_value,
                "reason": reason_value,
                "raw_notes": notes_value,
                "extracted_text": pasted_text,
            })
            save_day(selected_day, data)
            st.success("Weekly draft saved.")

        if b.button("Confirm Weekly Complete ✅", width="stretch"):
            weekly.update({
                "actual": actual,
                "expected": expected,
                "total_storage": total_storage,
                "five_year_avg": five_year_avg,
                "last_year": last_year,
                "surprise": actual - expected,
                "vs_5y": total_storage - five_year_avg,
                "vs_last_year": total_storage - last_year,
                "score": score_value,
                "bias": selected_bias,
                "summary": summary_value,
                "reason": reason_value,
                "raw_notes": notes_value,
                "extracted_text": pasted_text,
                "completed": True,
                "confirmed_at": datetime.now().isoformat(),
            })
            save_day(selected_day, data)
            st.success("Weekly storage completed.")
            st.rerun()

    if weekly.get("completed"):
        st.success("Weekly Storage is completed.")
    else:
        st.warning("Weekly Storage is not completed yet.")


with tabs[-2]:
    ensure_monthly_steo(data)
    monthly = data["monthly"][MONTHLY_MODULE]

    st.subheader("Monthly STEO")
    st.caption("Monthly structural regime layer. Use STEO Excel/text/PDF narrative as background context, not trade trigger.")

    left, right = st.columns([1, 1])

    with left:
        st.markdown("### Source upload")
        steo_xlsx = st.file_uploader(
            "Upload STEO_m.xlsx / STEO data workbook",
            type=["xlsx", "xlsm", "xls"],
            key="upload_monthly_steo_xlsx",
        )
        if steo_xlsx is not None:
            monthly["source_path"] = save_uploaded_source(selected_day, "monthly_steo_data", steo_xlsx)
            save_day(selected_day, data)
            st.success(f"Saved STEO workbook: {monthly['source_path']}")

        steo_pdf_or_text = st.file_uploader(
            "Optional: upload STEO text/full PDF for archive",
            type=["pdf", "txt"],
            key="upload_monthly_steo_text",
        )
        if steo_pdf_or_text is not None:
            monthly["text_source_path"] = save_uploaded_source(selected_day, "monthly_steo_text", steo_pdf_or_text)
            save_day(selected_day, data)
            st.success(f"Saved STEO text/PDF: {monthly['text_source_path']}")

        st.markdown("### Narrative fallback")
        st.caption("Paste the Natural Gas / Weather STEO paragraphs here if workbook scan is not enough.")
        pasted_steo_text = st.text_area(
            "Paste STEO Natural Gas / Weather narrative",
            value=monthly.get("raw_notes", ""),
            height=240,
            key=f"monthly_paste_{monthly.get('field_version',0)}",
        )

        if monthly.get("source_path"):
            st.info(f"Workbook source: {monthly['source_path']}")
        if monthly.get("text_source_path"):
            st.info(f"Archive source: {monthly['text_source_path']}")

    with right:
        st.markdown("### Monthly regime extraction")

        c1, c2 = st.columns(2)
        with c1:
            monthly["release_month"] = st.text_input("STEO release month", value=monthly.get("release_month", ""))
            st.caption("Auto-filled from STEO fetch when available. Optional metadata; not used in score.")
        with c2:
            st.metric("Current monthly score", monthly.get("score", 0))
            st.metric("Current monthly bias", monthly.get("bias", "neutral"))

        if st.button("Fetch Latest STEO from EIA (no PDF / no XLSX)", width="stretch"):
            try:
                res = analyze_latest_steo_fetch()
                st.session_state["monthly_preview"] = res
                st.success("Latest STEO fetched and monthly preview generated.")
            except Exception as e:
                st.error(f"STEO fetch failed: {e}")

        if st.button("Extract Monthly STEO from Workbook Tables", width="stretch"):
            try:
                if not monthly.get("source_path"):
                    raise RuntimeError("Upload STEO workbook first.")
                res = extract_steo_from_xlsx(ROOT / monthly["source_path"])
                st.session_state["monthly_preview"] = res
                st.success("Monthly STEO workbook preview generated.")
            except Exception as e:
                st.error(f"Monthly STEO workbook extraction failed: {e}")

        if st.button("Analyze Monthly STEO from Pasted Text", width="stretch"):
            try:
                if not pasted_steo_text.strip():
                    raise RuntimeError("Paste STEO Natural Gas / Weather text first.")
                res = extract_steo_from_pdf_text(pasted_steo_text)
                res["source"] = "pasted_text"
                st.session_state["monthly_preview"] = res
                st.success("Monthly STEO text preview generated.")
            except Exception as e:
                st.error(f"Monthly STEO text analysis failed: {e}")

        if "monthly_preview" in st.session_state:
            st.json(st.session_state["monthly_preview"])
            if st.button("Apply Monthly STEO to Fields", width="stretch"):
                res = st.session_state["monthly_preview"]
                monthly.update({
                    "score": res.get("score", 0),
                    "bias": res.get("bias", "neutral"),
                    "summary": res.get("summary", ""),
                    "reason": res.get("reason", ""),
                    "metrics": res.get("metrics", {}),
                    "auto_result": res,
                    "raw_notes": pasted_steo_text or res.get("extracted_text", ""),
                    "release_month": monthly.get("release_month") or steo_release_label_from_result(res),
                    "field_version": int(monthly.get("field_version", 0)) + 1,
                })
                save_day(selected_day, data)
                del st.session_state["monthly_preview"]
                st.success("Monthly STEO applied.")
                st.rerun()

        st.divider()
        st.markdown("### Manual monthly review / override")
        fv = int(monthly.get("field_version", 0))
        manual_monthly = st.toggle("Manual monthly edit mode", value=True, key=f"manual_monthly_{fv}")
        disabled_monthly = not manual_monthly

        score_value = st.slider(
            "Monthly STEO Score (-2 bearish structural to +2 bullish structural)",
            -2, 2, int(monthly.get("score", 0)),
            key=f"monthly_score_{fv}",
            disabled=disabled_monthly,
        )
        bias_value = monthly.get("bias", "neutral") if monthly.get("bias", "neutral") in BIAS_OPTIONS else "neutral"
        selected_bias = st.selectbox(
            "Monthly STEO Bias",
            BIAS_OPTIONS,
            index=BIAS_OPTIONS.index(bias_value),
            key=f"monthly_bias_{fv}",
            disabled=disabled_monthly,
        )
        summary_value = st.text_area("Monthly summary", value=monthly.get("summary", ""), height=100, key=f"monthly_summary_{fv}", disabled=disabled_monthly)
        reason_value = st.text_area("Monthly reason", value=monthly.get("reason", ""), height=140, key=f"monthly_reason_{fv}", disabled=disabled_monthly)
        notes_value = st.text_area("Monthly raw notes", value=monthly.get("raw_notes", ""), height=100, key=f"monthly_notes_{fv}", disabled=disabled_monthly)

        a, b = st.columns(2)
        if a.button("Save Monthly Draft", width="stretch"):
            monthly.update({
                "score": score_value,
                "bias": selected_bias,
                "summary": summary_value,
                "reason": reason_value,
                "raw_notes": notes_value,
            })
            save_day(selected_day, data)
            st.success("Monthly STEO draft saved.")

        if b.button("Confirm Monthly Complete ✅", width="stretch"):
            monthly.update({
                "score": score_value,
                "bias": selected_bias,
                "summary": summary_value,
                "reason": reason_value,
                "raw_notes": notes_value,
                "completed": True,
                "confirmed_at": datetime.now().isoformat(),
            })
            save_day(selected_day, data)
            st.success("Monthly STEO completed.")
            st.rerun()

    if monthly.get("completed"):
        st.success("Monthly STEO is completed.")
    else:
        st.warning("Monthly STEO is not completed yet.")


with tabs[-1]:
    st.subheader("Daily Report Center")
    if st.button("Generate Report Now", width="stretch"):
        data["daily_report"] = build_report(data); save_day(selected_day, data); st.success("Report generated.")
    data["daily_report"] = st.text_area("Daily report markdown", value=data.get("daily_report",""), height=500)
    c1,c2 = st.columns(2)
    if c1.button("Save Report", width="stretch"):
        save_day(selected_day, data); st.success("Saved report.")
    if c2.button("Export .md", width="stretch"):
        rp = REPORT_DIR / f"{selected_day}_daily_report.md"; rp.write_text(data["daily_report"], encoding="utf-8"); save_day(selected_day, data); st.success(f"Exported: {rp.relative_to(ROOT)}")

    c3, c4 = st.columns(2)
    if c3.button("Export .txt", width="stretch"):
        txtp = REPORT_DIR / f"{selected_day}_daily_report.txt"; txtp.write_text(markdown_to_txt(data["daily_report"]), encoding="utf-8"); save_day(selected_day, data); st.success(f"Exported: {txtp.relative_to(ROOT)}")

    rp = REPORT_DIR / f"{selected_day}_daily_report.md"
    txtp = REPORT_DIR / f"{selected_day}_daily_report.txt"
    if rp.exists():
        st.download_button("Download Daily Report .md", rp.read_text(encoding="utf-8"), file_name=rp.name, mime="text/markdown", width="stretch")
    if txtp.exists():
        st.download_button("Download Daily Report .txt", txtp.read_text(encoding="utf-8"), file_name=txtp.name, mime="text/plain", width="stretch")
